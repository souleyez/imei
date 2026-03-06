#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from flask import Flask, request, jsonify, send_from_directory, redirect
from flask_cors import CORS
import sqlite3
import json
import os
from datetime import datetime

try:
    import openpyxl
    HAS_OPENPYXL = True
except:
    HAS_OPENPYXL = False

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

APP_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(APP_DIR, 'iot.db')

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    conn = get_db()
    c = conn.cursor()
    
    c.execute('''CREATE TABLE IF NOT EXISTS label_projects (
        id INTEGER PRIMARY KEY,
        name TEXT,
        imei_prefix TEXT,
        mac_prefix TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS labels (
        id INTEGER PRIMARY KEY,
        project_id INTEGER,
        imei TEXT,
        mac TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    
    c.execute('''CREATE TABLE IF NOT EXISTS operations (
        id INTEGER PRIMARY KEY,
        operation_type TEXT,
        table_name TEXT,
        record_id INTEGER,
        data_before TEXT,
        data_after TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # 批次表
    c.execute('''CREATE TABLE IF NOT EXISTS batches (
        id INTEGER PRIMARY KEY,
        project_id INTEGER,
        name TEXT,
        info TEXT,
        count INTEGER,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # 已生成标签表
    c.execute('''CREATE TABLE IF NOT EXISTS printed_labels (
        id INTEGER PRIMARY KEY,
        batch_id INTEGER,
        imei TEXT,
        mac TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    
    # 批次标签表
    c.execute('''CREATE TABLE IF NOT EXISTS batch_labels (
        id INTEGER PRIMARY KEY,
        batch_id INTEGER,
        imei TEXT,
        mac TEXT,
        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )''')
    
    conn.commit()
    conn.close()

def log_op(conn, op_type, table, record_id, data=None):
    c = conn.cursor()
    c.execute('INSERT INTO operations (operation_type, table_name, record_id, data_after) VALUES (?, ?, ?, ?)',
              (op_type, table, record_id, json.dumps(data) if data else None))

# 首页
@app.route('/')
def index():
    return send_from_directory(APP_DIR, 'index.html')

# 打印页面
@app.route('/print')
def print_page():
    return send_from_directory(APP_DIR, 'print.html')

# 项目页面
@app.route('/project')
def project_page():
    return send_from_directory(APP_DIR, 'project.html')

# 项目API
@app.route('/api/projects', methods=['GET'])
def get_projects():
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM label_projects ORDER BY id DESC')
    rows = c.fetchall()
    result = []
    for row in rows:
        d = dict(row)
        if d.get('exclude_data'):
            try:
                exclude = json.loads(d['exclude_data'])
                d['excludeCount'] = len(exclude)
            except:
                d['excludeCount'] = 0
        else:
            d['excludeCount'] = 0
        result.append(d)
    conn.close()
    return jsonify(result)

@app.route('/api/projects/<int:project_id>', methods=['GET'])
def get_project(project_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM label_projects WHERE id = ?', (project_id,))
    row = c.fetchone()
    if not row:
        return jsonify({'error': 'Not found'}), 404
    d = dict(row)
    if d.get('exclude_data'):
        try:
            exclude = json.loads(d['exclude_data'])
            d['excludeCount'] = len(exclude)
        except:
            d['excludeCount'] = 0
    else:
        d['excludeCount'] = 0
    conn.close()
    return jsonify(d)

@app.route('/api/projects', methods=['POST'])
def create_project():
    data = request.json
    conn = get_db()
    c = conn.cursor()
    exclude_data = json.dumps(data.get('excludeData', [])) if data.get('excludeData') else None
    c.execute('INSERT INTO label_projects (name, imei_prefix, mac_prefix, exclude_data) VALUES (?, ?, ?, ?)',
              (data['name'], data['imeiPrefix'], data['macPrefix'], exclude_data))
    pid = c.lastrowid
    log_op(conn, 'INSERT', 'label_projects', pid, data)
    conn.commit()
    conn.close()
    return jsonify({'id': pid})

@app.route('/api/projects/<int:project_id>', methods=['PUT'])
def update_project(project_id):
    data = request.json
    conn = get_db()
    c = conn.cursor()
    exclude_data = json.dumps(data.get('excludeData', [])) if data.get('excludeData') else None
    c.execute('UPDATE label_projects SET name=?, imei_prefix=?, mac_prefix=?, exclude_data=?, exclude_imei_start=?, exclude_imei_end=? WHERE id=?',
              (data['name'], data['imeiPrefix'], data['macPrefix'], exclude_data, 
               data.get('excludeImeiStart'), data.get('excludeImeiEnd'), project_id))
    log_op(conn, 'UPDATE', 'label_projects', project_id, data)
    conn.commit()
    conn.close()
    return jsonify({'success': True})

@app.route('/api/projects/<int:project_id>', methods=['DELETE'])
def delete_project(project_id):
    conn = get_db()
    c = conn.cursor()
    # 删除该项目的所有批次标签
    c.execute('DELETE FROM batch_labels WHERE batch_id IN (SELECT id FROM batches WHERE project_id=?)', (project_id,))
    # 删除该项目的所有批次
    c.execute('DELETE FROM batches WHERE project_id=?', (project_id,))
    # 删除项目
    c.execute('DELETE FROM label_projects WHERE id=?', (project_id,))
    log_op(conn, 'DELETE', 'label_projects', project_id, {'project_id': project_id})
    conn.commit()
    conn.close()
    return jsonify({'success': True})

# 批次API
@app.route('/api/batches', methods=['GET'])
def get_batches():
    project_id = request.args.get('project')
    conn = get_db()
    c = conn.cursor()
    if project_id:
        c.execute('SELECT * FROM batches WHERE project_id = ? ORDER BY id DESC', (project_id,))
    else:
        c.execute('SELECT * FROM batches ORDER BY id DESC')
    rows = c.fetchall()
    result = [dict(row) for row in rows]
    conn.close()
    return jsonify(result)

# 解析排除表文件
@app.route('/api/parse-exclude', methods=['POST'])
def parse_exclude():
    if 'file' not in request.files:
        return jsonify([])
    
    file = request.files['file']
    filename = file.filename.lower()
    result = []
    
    try:
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            if not HAS_OPENPYXL:
                return jsonify([])
            import io
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            for row in ws.iter_rows(values_only=True):
                if not row or not row[0]:
                    continue
                imei = str(row[0]).strip() if row[0] else ''
                mac = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                # 格式化MAC（12位转冒号分隔）
                if mac and len(mac) == 12:
                    mac = ':'.join([mac[i:i+2] for i in range(0,12,2)]).upper()
                if imei and imei.lower() != 'imei':
                    result.append({'imei': imei, 'mac': mac})
        else:
            # CSV/TXT
            content = file.read().decode('utf-8', errors='ignore')
            for line in content.split('\n'):
                parts = line.strip().split(',')
                if parts and parts[0] and parts[0].lower() != 'imei':
                    imei = parts[0].strip()
                    mac = parts[1].strip() if len(parts) > 1 else ''
                    result.append({'imei': imei, 'mac': mac})
    except Exception as e:
        print('Parse error:', e)
        return jsonify([])
    
    return jsonify(result)

@app.route('/api/batches', methods=['POST'])
def create_batch():
    data = request.json
    conn = get_db()
    c = conn.cursor()
    c.execute('INSERT INTO batches (project_id, name, info, count) VALUES (?, ?, ?, ?)',
              (data['project_id'], data['name'], data['info'], len(data['imeis'])))
    batch_id = c.lastrowid
    
    for imei, mac in zip(data['imeis'], data['macs']):
        c.execute('INSERT INTO batch_labels (batch_id, imei, mac) VALUES (?, ?, ?)',
                  (batch_id, imei, mac))
        log_op(conn, 'INSERT', 'batch_labels', c.lastrowid, {'imei': imei, 'mac': mac})
    
    conn.commit()
    conn.close()
    return jsonify({'id': batch_id})

@app.route('/api/batches/<int:batch_id>', methods=['PUT'])
def update_batch(batch_id):
    data = request.json
    conn = get_db()
    c = conn.cursor()
    c.execute('UPDATE batches SET name = ?, info = ? WHERE id = ?',
              (data.get('name', ''), data.get('info', ''), batch_id))
    log_op(conn, 'UPDATE', 'batches', batch_id, data)
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/batches/<int:batch_id>', methods=['DELETE'])
def delete_batch(batch_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('DELETE FROM batch_labels WHERE batch_id = ?', (batch_id,))
    c.execute('DELETE FROM batches WHERE id = ?', (batch_id,))
    log_op(conn, 'DELETE', 'batches', batch_id, {'batch_id': batch_id})
    conn.commit()
    conn.close()
    return jsonify({'ok': True})

@app.route('/api/batches/<int:batch_id>/groups', methods=['GET'])
def get_batch_groups(batch_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM batches WHERE id = ?', (batch_id,))
    batch = dict(c.fetchone())
    c.execute('SELECT * FROM batch_labels WHERE batch_id = ? ORDER BY id', (batch_id,))
    labels = c.fetchall()
    conn.close()
    
    groups = []
    for i in range(0, len(labels), 12):
        group_labels = labels[i:i+12]
        imeis = [l['imei'] for l in group_labels]
        macs = [l['mac'] for l in group_labels]
        combined = imeis + macs
        groups.append({'imeis': imeis, 'macs': macs, 'combined': ','.join(combined)})
    
    return jsonify({'batch': batch, 'groups': groups})

@app.route('/api/batches/<int:batch_id>/download', methods=['GET'])
def download_batch(batch_id):
    from urllib.parse import quote
    import io
    from openpyxl import Workbook
    
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM batches WHERE id = ?', (batch_id,))
    batch = c.fetchone()
    c.execute('SELECT * FROM batch_labels WHERE batch_id = ? ORDER BY id', (batch_id,))
    labels = c.fetchall()
    conn.close()
    
    # 创建xlsx
    wb = Workbook()
    ws = wb.active
    ws.title = '标签数据'
    ws.append(['IMEI', 'MAC', '文字信息'])
    
    info = batch['info'] or ''
    for l in labels:
        mac_no_colon = l['mac'].replace(':', '') if l['mac'] else ''
        ws.append([l['imei'], mac_no_colon, info])
    
    filename = batch['name'] or 'batch'
    safe_filename = quote(filename) + '.xlsx'
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.read(), 200, {'Content-Type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 'Content-Disposition': f'attachment; filename="{safe_filename}"'}

# 获取批次标签
@app.route('/api/batches/<int:batch_id>/labels', methods=['GET'])
def get_batch_labels(batch_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT imei, mac FROM batch_labels WHERE batch_id = ? ORDER BY id', (batch_id,))
    labels = c.fetchall()
    conn.close()
    return jsonify([{'imei': l['imei'], 'mac': l['mac']} for l in labels])

# 操作记录
@app.route('/api/operations', methods=['GET'])
def get_ops():
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM operations ORDER BY id DESC LIMIT 100')
    rows = c.fetchall()
    conn.close()
    return jsonify([dict(row) for row in rows])

# 项目标签（兼容）
@app.route('/api/projects/<int:project_id>/box-labels', methods=['GET'])
def get_project_box_labels(project_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM label_projects WHERE id = ?', (project_id,))
    project = c.fetchone()
    if not project:
        return jsonify({'error': 'not found'})
    c.execute('SELECT * FROM labels WHERE project_id = ? ORDER BY id', (project_id,))
    labels = c.fetchall()
    conn.close()
    
    groups = []
    for i in range(0, len(labels), 12):
        gl = labels[i:i+12]
        imeis = [l['imei'] for l in gl]
        macs = [l['mac'] for l in gl]
        groups.append({
            'group_no': len(groups)+1,
            'imeis': imeis,
            'macs': macs,
            'combined': ','.join(imeis + macs)
        })
    
    return jsonify({'project_name': project['name'], 'groups': groups})

# 导出Excel
@app.route('/api/projects/<int:project_id>/export-excel', methods=['GET'])
def export_excel(project_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('SELECT * FROM labels WHERE project_id = ? ORDER BY id', (project_id,))
    labels = c.fetchall()
    conn.close()
    
    csv = 'IMEI,MAC\n'
    for l in labels:
        csv += f"{l['imei']},{l['mac']}\n"
    return csv, 200, {'Content-Type': 'text/csv', 'Content-Disposition': 'attachment; filename=labels.csv'}

# 保存已打印标签
@app.route('/api/printed-labels', methods=['POST'])
def save_printed_labels():
    data = request.json
    conn = get_db()
    c = conn.cursor()
    batch_id = data.get('batch_id')
    imeis = data.get('imeis', [])
    macs = data.get('macs', [])
    
    for i in range(len(imeis)):
        c.execute('INSERT INTO printed_labels (batch_id, imei, mac) VALUES (?, ?, ?)',
                  (batch_id, imeis[i], macs[i] if i < len(macs) else ''))
    
    log_op(conn, 'INSERT', 'printed_labels', batch_id, {'batch_id': batch_id, 'count': len(imeis)})
    conn.commit()
    conn.close()
    return jsonify({'success': True, 'count': len(imeis)})

# 获取已打印标签
@app.route('/api/printed-labels', methods=['GET'])
def get_printed_labels():
    batch_id = request.args.get('batch')
    conn = get_db()
    c = conn.cursor()
    if batch_id:
        c.execute('SELECT * FROM printed_labels WHERE batch_id = ? ORDER BY id', (batch_id,))
    else:
        c.execute('SELECT pl.*, b.name as batch_name FROM printed_labels pl LEFT JOIN batches b ON pl.batch_id = b.id ORDER BY pl.id DESC')
    labels = c.fetchall()
    conn.close()
    return jsonify(labels)

# 删除已打印标签
@app.route('/api/printed-labels/<int:label_id>', methods=['DELETE'])
def delete_printed_label(label_id):
    conn = get_db()
    c = conn.cursor()
    c.execute('DELETE FROM printed_labels WHERE id = ?', (label_id,))
    conn.commit()
    conn.close()
    return jsonify({'success': True})


# ==================== KnowledgeHub 代理 ====================
from urllib.parse import urlparse, parse_qs

@app.route('/how')
def proxy_knowledgehub_root():
    from flask import redirect
    return redirect('/how/', code=302)

@app.route('/how/', defaults={'path': ''})
@app.route('/how/<path:path>')
def proxy_knowledgehub(path):
    """代理到 KnowledgeHub 应用 (端口 5000)"""
    import requests
    target_url = 'http://127.0.0.1:5000/' + path
    
    # 处理query string
    qs = request.query_string.decode()
    if qs:
        # 如果有 next 参数且指向 / 开头，转换为 /how/ 开头
        if 'next=' in qs:
            qs = qs.replace('next=%2F', 'next=%2Fhow%2F')
        target_url += '?' + qs
    
    headers = {k: v for k, v in request.headers if k.lower() not in ('host', 'cookie')}
    if request.cookies:
        headers['Cookie'] = '; '.join([f'{k}={v}' for k, v in request.cookies.items()])
    
    try:
        resp = requests.request(
            method=request.method,
            url=target_url,
            headers=headers,
            data=request.get_data(),
            cookies=request.cookies,
            allow_redirects=False
        )
        
        # 处理重定向
        new_headers = dict(resp.headers)
        loc = new_headers.get('Location') or new_headers.get('location')
        if loc and not loc.startswith('http'):
            new_loc = '/how/' + loc.lstrip('/')
            new_headers['Location'] = new_loc
        
        from flask import Response
        return Response(
            resp.content,
            status=resp.status_code,
            headers=new_headers
        )
    except Exception as e:
        return '代理错误: ' + str(e), 502

if __name__ == '__main__':
    init_db()
    print("IOT工具集 http://0.0.0.0:5001")
    app.run(host='0.0.0.0', port=5001, debug=False)
